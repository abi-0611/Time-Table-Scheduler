from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REF_DIR = ROOT / "Reference Data"
OUT_MD = ROOT / "docs" / "reference_spreadsheet_analysis.md"

DAY_TOKENS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", s).strip()


def _sheet_bbox(ws) -> tuple[int, int, int, int]:
    """Return min_row, max_row, min_col, max_col of non-empty cells."""

    min_r = None
    min_c = None
    max_r = 0
    max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if _norm(cell.value) == "":
                continue
            r, c = cell.row, cell.column
            min_r = r if min_r is None else min(min_r, r)
            min_c = c if min_c is None else min(min_c, c)
            max_r = max(max_r, r)
            max_c = max(max_c, c)
    if min_r is None or min_c is None:
        return (1, 1, 1, 1)
    return (min_r, max_r, min_c, max_c)


def _grid_preview(ws, *, min_row: int, max_row: int, min_col: int, max_col: int, rows: int = 18, cols: int = 14) -> list[list[str]]:
    r2 = min(max_row, min_row + rows - 1)
    c2 = min(max_col, min_col + cols - 1)
    out: list[list[str]] = []
    for r in range(min_row, r2 + 1):
        row_vals: list[str] = []
        for c in range(min_col, c2 + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if len(v) > 60:
                v = v[:57] + "…"
            row_vals.append(v)
        out.append(row_vals)
    return out


def _markdown_table(grid: list[list[str]]) -> str:
    if not grid:
        return ""
    ncols = max(len(r) for r in grid)
    header = [f"C{i+1}" for i in range(ncols)]

    def esc(s: str) -> str:
        return str(s).replace("\n", " ").replace("|", "\\|")

    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * ncols) + " |")
    for r in grid:
        r = r + [""] * (ncols - len(r))
        lines.append("| " + " | ".join(esc(x) for x in r) + " |")
    return "\n".join(lines)


def _find_day_like_rows(grid: list[list[str]]) -> list[int]:
    hits = []
    for i, row in enumerate(grid):
        joined = " ".join(row).lower()
        if sum(tok in joined for tok in DAY_TOKENS) >= 2:
            hits.append(i)
    return hits


def _collect_unique_strings(ws, *, min_row: int, max_row: int, min_col: int, max_col: int, limit: int = 2000) -> list[str]:
    uniq: set[str] = set()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if not v:
                continue
            if v.isdigit():
                continue
            uniq.add(v)
            if len(uniq) >= limit:
                return sorted(uniq)
    return sorted(uniq)


def analyze_workbook(path: Path) -> str:
    wb = load_workbook(path, data_only=False)
    md: list[str] = []
    md.append(f"# {path.name}")
    md.append("")
    md.append(f"Sheets: {len(wb.sheetnames)}")
    md.append("")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        min_r, max_r, min_c, max_c = _sheet_bbox(ws)
        merged = list(ws.merged_cells.ranges)
        md.append(f"## Sheet: {sheet_name}")
        md.append("")
        md.append(f"- Used range (detected): rows {min_r}–{max_r}, cols {min_c}–{max_c}")
        md.append(f"- Merged ranges: {len(merged)}")
        if merged:
            md.append("- Example merged ranges: " + ", ".join(str(rng) for rng in merged[:8]) + ("" if len(merged) <= 8 else " …"))
        md.append("")

        grid = _grid_preview(ws, min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c)
        md.append("Preview (top-left):")
        md.append("")
        md.append(_markdown_table(grid))
        md.append("")

        day_rows = _find_day_like_rows(grid)
        if day_rows:
            md.append(f"- Day-like rows in preview (0-based within preview): {day_rows}")
        else:
            md.append("- Day-like rows in preview: none detected")

        uniq = _collect_unique_strings(ws, min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, limit=400)
        md.append("")
        md.append("Distinct non-empty strings (sample):")
        md.append("")
        if uniq:
            md.append("- " + "\n- ".join(uniq[:80]) + ("" if len(uniq) <= 80 else "\n- …"))
        else:
            md.append("- (none)")

        md.append("")

    return "\n".join(md).strip() + "\n"


def main() -> None:
    if not REF_DIR.exists():
        raise SystemExit(f"Reference Data folder not found: {REF_DIR}")

    files = [
        REF_DIR / "Class TT 25-26 Odd AI & DS.xlsx",
        REF_DIR / "Individual_TT_25-26_Odd (2).xlsx",
        REF_DIR / "AI&DS Staff WL 2025-26 Even.xlsx",
        REF_DIR / "AI-DS Work Load 2025-26 Even.xlsx",
    ]

    missing = [str(p) for p in files if not p.exists()]
    if missing:
        raise SystemExit("Missing files:\n" + "\n".join(missing))

    sections: list[str] = []
    sections.append("# Reference Spreadsheet Analysis (Auto-Extracted)")
    sections.append("")
    sections.append(
        "This document is generated by `scripts/analyze_reference_spreadsheets.py`. "
        "It extracts sheet-by-sheet previews and string samples to support the Phase 1 analysis."
    )
    sections.append("")

    for p in files:
        sections.append(analyze_workbook(p))
        sections.append("\n---\n")

    OUT_MD.parent.mkdir(parents=True, exist_ok=True)
    OUT_MD.write_text("\n".join(sections).strip() + "\n", encoding="utf-8")
    print(f"Wrote: {OUT_MD}")


if __name__ == "__main__":
    main()
